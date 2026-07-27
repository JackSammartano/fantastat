"""Ispezione read-only dei workbook storici Fantacalcio.

Il modulo produce un report deterministico e non modifica mai i file sorgente.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import openpyxl


NUMERIC_COLUMNS = (
    "Pv",
    "Mv",
    "Fm",
    "Gf",
    "Gs",
    "Rp",
    "Rc",
    "R+",
    "R-",
    "Ass",
    "Amm",
    "Esp",
    "Au",
)
AVERAGE_COLUMNS = ("Mv", "Fm")


@dataclass(frozen=True)
class InspectionConfig:
    """Configurazione risolta per l'ispezione."""

    source_root: Path
    canonical_sheet: str
    header_row: int
    role_sheets: dict[str, str]
    expected_columns: tuple[str, ...]
    seasons: tuple[dict[str, str], ...]


def sha256_file(path: Path) -> str:
    """Calcola l'hash del file senza alterarlo."""

    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_config(config_path: Path) -> InspectionConfig:
    """Carica e valida la configurazione YAML in sintassi JSON compatibile."""

    with config_path.open("r", encoding="utf-8") as stream:
        raw = json.load(stream)

    required = {
        "source_root",
        "canonical_sheet",
        "header_row",
        "role_sheets",
        "expected_columns",
        "seasons",
    }
    missing = sorted(required - set(raw or {}))
    if missing:
        raise ValueError(f"Configurazione incompleta; chiavi mancanti: {missing}")

    source_root = (config_path.parent / raw["source_root"]).resolve()
    seasons = tuple(raw["seasons"])
    codes = [season["code"] for season in seasons]
    if len(codes) != len(set(codes)):
        raise ValueError("I codici stagione devono essere univoci")

    return InspectionConfig(
        source_root=source_root,
        canonical_sheet=str(raw["canonical_sheet"]),
        header_row=int(raw["header_row"]),
        role_sheets={
            str(sheet): str(role) for sheet, role in raw["role_sheets"].items()
        },
        expected_columns=tuple(map(str, raw["expected_columns"])),
        seasons=seasons,
    )


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _headers(worksheet: openpyxl.worksheet.worksheet.Worksheet, row: int) -> list[str]:
    return [
        "" if cell.value is None else str(cell.value).strip()
        for cell in worksheet[row]
    ]


def _records(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    headers: Sequence[str],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for values in worksheet.iter_rows(
        min_row=header_row + 1,
        max_col=len(headers),
        values_only=True,
    ):
        if all(value is None for value in values):
            continue
        records.append(
            {
                header: _json_value(value)
                for header, value in zip(headers, values, strict=True)
            }
        )
    return records


def _type_name(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    return "string"


def _duplicate_count(records: Sequence[Mapping[str, Any]], fields: Sequence[str]) -> int:
    keys = [tuple(record.get(field) for field in fields) for record in records]
    counts = Counter(keys)
    return sum(count - 1 for count in counts.values() if count > 1)


def _row_key(record: Mapping[str, Any], headers: Sequence[str]) -> str:
    return json.dumps(
        [record.get(header) for header in headers],
        ensure_ascii=False,
        separators=(",", ":"),
    )


def inspect_sheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Ispeziona un foglio e restituisce metadati e record."""

    headers = _headers(worksheet, header_row)
    records = _records(worksheet, header_row, headers)
    type_sets = {
        header: sorted({_type_name(record.get(header)) for record in records})
        for header in headers
    }
    missing_values = {
        header: sum(record.get(header) is None for record in records)
        for header in headers
    }
    empty_columns = [
        header for header, count in missing_values.items() if count == len(records)
    ]
    formulas = sorted(
        cell.coordinate
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )

    summary = {
        "name": worksheet.title,
        "rows_including_headers": worksheet.max_row,
        "data_rows": len(records),
        "columns": len(headers),
        "headers": headers,
        "detected_types": type_sets,
        "missing_values": missing_values,
        "empty_columns": empty_columns,
        "exact_duplicate_rows": _duplicate_count(records, headers),
        "duplicate_ids": (
            _duplicate_count(records, ("Id",)) if "Id" in headers else None
        ),
        "merged_cells": sorted(map(str, worksheet.merged_cells.ranges)),
        "formula_count": len(formulas),
        "formula_cells": formulas,
        "sample_rows": records[:5],
    }
    return summary, records


def _numeric_summary(records: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for column in NUMERIC_COLUMNS:
        values = [
            record[column]
            for record in records
            if isinstance(record.get(column), (int, float))
            and not isinstance(record.get(column), bool)
        ]
        result[column] = {
            "minimum": min(values) if values else None,
            "maximum": max(values) if values else None,
            "negative_count": sum(value < 0 for value in values),
        }
    return result


def _zero_pv_anomalies(records: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    anomalies = []
    for record in records:
        if record.get("Pv") != 0:
            continue
        non_zero = [
            column
            for column in NUMERIC_COLUMNS
            if column not in ("Pv", *AVERAGE_COLUMNS)
            and isinstance(record.get(column), (int, float))
            and record[column] != 0
        ]
        if non_zero:
            anomalies.append(
                {
                    "Id": record.get("Id"),
                    "Nome": record.get("Nome"),
                    "Squadra": record.get("Squadra"),
                    "non_zero_columns": non_zero,
                }
            )
    return anomalies


def inspect_workbook(
    path: Path,
    season_code: str,
    config: InspectionConfig,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Ispeziona un workbook configurato."""

    if not path.is_file():
        raise FileNotFoundError(f"File stagione non trovato: {path}")

    workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    try:
        sheets: dict[str, Any] = {}
        records_by_sheet: dict[str, list[dict[str, Any]]] = {}
        for worksheet in workbook.worksheets:
            summary, records = inspect_sheet(worksheet, config.header_row)
            sheets[worksheet.title] = summary
            records_by_sheet[worksheet.title] = records

        if config.canonical_sheet not in records_by_sheet:
            raise ValueError(
                f"Foglio canonico {config.canonical_sheet!r} assente in {path.name}"
            )

        canonical = records_by_sheet[config.canonical_sheet]
        canonical_headers = sheets[config.canonical_sheet]["headers"]
        header_issues = {
            "missing": sorted(set(config.expected_columns) - set(canonical_headers)),
            "unexpected": sorted(set(canonical_headers) - set(config.expected_columns)),
            "order_matches": canonical_headers == list(config.expected_columns),
        }

        role_checks: dict[str, Any] = {}
        for sheet_name, role in config.role_sheets.items():
            if sheet_name not in records_by_sheet:
                role_checks[sheet_name] = {"status": "missing"}
                continue
            sheet_records = records_by_sheet[sheet_name]
            expected_ids = {
                record.get("Id")
                for record in canonical
                if record.get("R") == role
            }
            actual_ids = {record.get("Id") for record in sheet_records}
            role_checks[sheet_name] = {
                "status": "ok" if expected_ids == actual_ids else "mismatch",
                "expected_role": role,
                "data_rows": len(sheet_records),
                "invalid_role_rows": sum(
                    record.get("R") != role for record in sheet_records
                ),
                "missing_ids": sorted(expected_ids - actual_ids),
                "unexpected_ids": sorted(actual_ids - expected_ids),
            }

        duplicate_names = _duplicate_count(canonical, ("Nome",))
        zero_pv = sum(record.get("Pv") == 0 for record in canonical)
        zero_pv_with_nonzero_stats = _zero_pv_anomalies(canonical)
        penalty_mismatches = [
            record.get("Id")
            for record in canonical
            if all(
                isinstance(record.get(column), (int, float))
                for column in ("Rc", "R+", "R-")
            )
            and record["Rc"] != record["R+"] + record["R-"]
        ]
        report = {
            "season": season_code,
            "file": path.name,
            "sha256": sha256_file(path),
            "sheet_names": workbook.sheetnames,
            "sheets": sheets,
            "canonical_sheet": config.canonical_sheet,
            "canonical_data_rows": len(canonical),
            "header_issues": header_issues,
            "role_sheet_checks": role_checks,
            "canonical_duplicate_names": duplicate_names,
            "canonical_zero_pv_rows": zero_pv,
            "zero_pv_with_nonzero_stats": zero_pv_with_nonzero_stats,
            "penalty_identity_mismatch_ids": penalty_mismatches,
            "numeric_summary": _numeric_summary(canonical),
        }
        return report, canonical
    finally:
        workbook.close()


def cross_season_summary(
    season_records: Sequence[tuple[str, Sequence[Mapping[str, Any]]]],
) -> dict[str, Any]:
    """Rileva stabilità e collisioni delle identità fra stagioni."""

    by_id: dict[Any, list[dict[str, Any]]] = defaultdict(list)
    by_name: dict[Any, list[dict[str, Any]]] = defaultdict(list)
    total_records = 0

    for season, records in season_records:
        for source_record in records:
            record = dict(source_record)
            record["season"] = season
            by_id[record.get("Id")].append(record)
            by_name[record.get("Nome")].append(record)
            total_records += 1

    id_name_variants = []
    id_role_changes = []
    for external_id, records in sorted(by_id.items(), key=lambda item: item[0]):
        names = sorted({str(record.get("Nome")) for record in records})
        roles = sorted({str(record.get("R")) for record in records})
        seasons = sorted({str(record["season"]) for record in records})
        if len(names) > 1:
            id_name_variants.append(
                {"Id": external_id, "names": names, "seasons": seasons}
            )
        if len(roles) > 1:
            id_role_changes.append(
                {"Id": external_id, "roles": roles, "seasons": seasons}
            )

    name_id_collisions = []
    for name, records in sorted(by_name.items(), key=lambda item: str(item[0])):
        ids = sorted({record.get("Id") for record in records})
        if len(ids) > 1:
            name_id_collisions.append(
                {
                    "Nome": name,
                    "ids": ids,
                    "seasons": sorted({record["season"] for record in records}),
                }
            )

    season_availability = Counter(
        len({record["season"] for record in records}) for records in by_id.values()
    )
    return {
        "total_records": total_records,
        "unique_external_ids": len(by_id),
        "players_by_available_seasons": {
            str(count): season_availability[count]
            for count in sorted(season_availability)
        },
        "id_name_variant_count": len(id_name_variants),
        "id_name_variants": id_name_variants,
        "id_role_change_count": len(id_role_changes),
        "id_role_changes": id_role_changes,
        "name_id_collision_count": len(name_id_collisions),
        "name_id_collisions": name_id_collisions,
    }


def inspect_all(config: InspectionConfig) -> dict[str, Any]:
    """Esegue l'ispezione completa definita dalla configurazione."""

    workbook_reports = []
    season_records = []
    for season in config.seasons:
        path = config.source_root / season["file"]
        report, records = inspect_workbook(path, season["code"], config)
        workbook_reports.append(report)
        season_records.append((season["code"], records))

    return {
        "report_schema_version": 1,
        "source_root": ".",
        "canonical_sheet": config.canonical_sheet,
        "expected_columns": list(config.expected_columns),
        "workbooks": workbook_reports,
        "cross_season": cross_season_summary(season_records),
    }


def render_markdown(report: Mapping[str, Any]) -> str:
    """Rende un riepilogo Markdown stabile e leggibile."""

    lines = [
        "# Report di ispezione Excel",
        "",
        f"Schema report: `{report['report_schema_version']}`",
        "",
        "## Riepilogo workbook",
        "",
        "| Stagione | File | Righe | ID duplicati | Nomi duplicati | Pv=0 |",
        "|---|---|---:|---:|---:|---:|",
    ]
    for workbook in report["workbooks"]:
        canonical = workbook["sheets"][workbook["canonical_sheet"]]
        lines.append(
            "| {season} | `{file}` | {rows} | {ids} | {names} | {zero} |".format(
                season=workbook["season"],
                file=workbook["file"],
                rows=workbook["canonical_data_rows"],
                ids=canonical["duplicate_ids"],
                names=workbook["canonical_duplicate_names"],
                zero=workbook["canonical_zero_pv_rows"],
            )
        )

    cross = report["cross_season"]
    lines.extend(
        [
            "",
            "## Controlli cross-stagione",
            "",
            f"- Record complessivi: **{cross['total_records']}**.",
            f"- ID esterni distinti: **{cross['unique_external_ids']}**.",
            f"- ID con varianti del nome: **{cross['id_name_variant_count']}**.",
            f"- ID con cambi di ruolo: **{cross['id_role_change_count']}**.",
            f"- Nomi associati a ID differenti: **{cross['name_id_collision_count']}**.",
            "",
            "### Disponibilità per numero di stagioni",
            "",
        ]
    )
    for seasons, count in cross["players_by_available_seasons"].items():
        lines.append(f"- {seasons} stagione/i: {count} giocatori.")

    lines.extend(["", "### Collisioni nome–ID", ""])
    if cross["name_id_collisions"]:
        for collision in cross["name_id_collisions"]:
            lines.append(
                f"- `{collision['Nome']}`: ID {collision['ids']}, "
                f"stagioni {collision['seasons']}."
            )
    else:
        lines.append("- Nessuna collisione.")

    lines.extend(["", "## Anomalie per workbook", ""])
    for workbook in report["workbooks"]:
        lines.append(f"### {workbook['season']}")
        lines.append("")
        if workbook["zero_pv_with_nonzero_stats"]:
            for anomaly in workbook["zero_pv_with_nonzero_stats"]:
                lines.append(
                    f"- ID {anomaly['Id']} — {anomaly['Nome']}: `Pv=0`, "
                    f"valori non zero in {anomaly['non_zero_columns']}."
                )
        else:
            lines.append("- Nessuna statistica additiva non zero con `Pv=0`.")
        if workbook["penalty_identity_mismatch_ids"]:
            lines.append(
                "- Identità `Rc = R+ + R-` non rispettata per ID: "
                f"{workbook['penalty_identity_mismatch_ids']}."
            )
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def write_reports(report: Mapping[str, Any], output_dir: Path) -> tuple[Path, Path]:
    """Scrive i due report generati in una directory separata."""

    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "excel-inspection.json"
    markdown_path = output_dir / "excel-inspection.md"
    json_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    markdown_path.write_text(render_markdown(report), encoding="utf-8")
    return json_path, markdown_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Ispeziona in sola lettura i workbook storici Fantacalcio."
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=Path("backend/config/seasons.yaml"),
        help="Percorso della configurazione YAML.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("reports/data-quality"),
        help="Directory separata in cui salvare i report.",
    )
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    config = load_config(args.config.resolve())
    report = inspect_all(config)
    json_path, markdown_path = write_reports(report, args.output_dir.resolve())
    print(f"Report JSON: {json_path}")
    print(f"Report Markdown: {markdown_path}")
    print(f"Workbook analizzati: {len(report['workbooks'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
