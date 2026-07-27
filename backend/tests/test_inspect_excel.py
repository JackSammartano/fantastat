from __future__ import annotations

import hashlib
import json
from pathlib import Path

import openpyxl

from backend.scripts.inspect_excel import (
    InspectionConfig,
    cross_season_summary,
    inspect_workbook,
    render_markdown,
    write_reports,
)


HEADERS = (
    "Id",
    "R",
    "Rm",
    "Nome",
    "Squadra",
    "Pv",
    "Mv",
    "Fm",
    "Gf",
    "Gs",
    "Rp",
    "Rc",
    "R+",
    "R-",
    "Ass",
    "Amm",
    "Esp",
    "Au",
)
ROLE_SHEETS = {
    "Portieri": "P",
    "Difensori": "D",
    "Centrocampisti": "C",
    "Attaccanti": "A",
}


def _hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _row(
    external_id: int,
    role: str,
    name: str,
    *,
    pv: int = 1,
    yellow_cards: int = 0,
) -> list[object]:
    return [
        external_id,
        role,
        {"P": "Por", "D": "Dc", "C": "C", "A": "Pc"}[role],
        name,
        "Squadra Test",
        pv,
        6.0 if pv else 0.0,
        6.0 if pv else 0.0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        yellow_cards,
        0,
        0,
    ]


def _create_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for sheet_name in ("Tutti", *ROLE_SHEETS):
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.merge_cells("A1:R1")
        worksheet.cell(1, 1, "Statistiche test")
        worksheet.append(HEADERS)
        sheet_rows = (
            rows
            if sheet_name == "Tutti"
            else [row for row in rows if row[1] == ROLE_SHEETS[sheet_name]]
        )
        for row in sheet_rows:
            worksheet.append(row)
    workbook.save(path)


def _config(tmp_path: Path) -> InspectionConfig:
    return InspectionConfig(
        source_root=tmp_path,
        canonical_sheet="Tutti",
        header_row=2,
        role_sheets=ROLE_SHEETS,
        expected_columns=HEADERS,
        seasons=(),
    )


def test_inspection_does_not_modify_source_and_checks_role_sheets(
    tmp_path: Path,
) -> None:
    path = tmp_path / "season.xlsx"
    rows = [
        _row(1, "P", "Portiere"),
        _row(2, "D", "Difensore"),
        _row(3, "C", "Centrocampista"),
        _row(4, "A", "Attaccante"),
    ]
    _create_workbook(path, rows)
    original_hash = _hash(path)

    report, records = inspect_workbook(path, "2022/2023", _config(tmp_path))

    assert _hash(path) == original_hash
    assert len(records) == 4
    assert report["header_issues"] == {
        "missing": [],
        "unexpected": [],
        "order_matches": True,
    }
    assert all(
        check["status"] == "ok"
        for check in report["role_sheet_checks"].values()
    )
    assert report["sheets"]["Tutti"]["formula_count"] == 0
    assert report["sheets"]["Tutti"]["merged_cells"] == ["A1:R1"]


def test_inspection_reports_zero_pv_with_nonzero_additive_stat(
    tmp_path: Path,
) -> None:
    path = tmp_path / "season.xlsx"
    _create_workbook(path, [_row(5785, "A", "Lazetic", pv=0, yellow_cards=1)])

    report, _ = inspect_workbook(path, "2022/2023", _config(tmp_path))

    assert report["canonical_zero_pv_rows"] == 1
    assert report["zero_pv_with_nonzero_stats"] == [
        {
            "Id": 5785,
            "Nome": "Lazetic",
            "Squadra": "Squadra Test",
            "non_zero_columns": ["Amm"],
        }
    ]


def test_cross_season_summary_keeps_aliases_and_homonyms_separate() -> None:
    first = [
        dict(zip(HEADERS, _row(10, "C", "Soule'"), strict=True)),
        dict(zip(HEADERS, _row(20, "A", "Ndiaye"), strict=True)),
    ]
    second = [
        dict(zip(HEADERS, _row(10, "A", "Soulè"), strict=True)),
        dict(zip(HEADERS, _row(30, "A", "Ndiaye"), strict=True)),
    ]

    summary = cross_season_summary(
        [("2022/2023", first), ("2023/2024", second)]
    )

    assert summary["unique_external_ids"] == 3
    assert summary["id_name_variant_count"] == 1
    assert summary["id_role_change_count"] == 1
    assert summary["name_id_collision_count"] == 1
    assert summary["name_id_collisions"][0]["ids"] == [20, 30]


def test_reports_are_deterministic(tmp_path: Path) -> None:
    report = {
        "report_schema_version": 1,
        "workbooks": [],
        "cross_season": {
            "total_records": 0,
            "unique_external_ids": 0,
            "players_by_available_seasons": {},
            "id_name_variant_count": 0,
            "id_name_variants": [],
            "id_role_change_count": 0,
            "id_role_changes": [],
            "name_id_collision_count": 0,
            "name_id_collisions": [],
        },
    }
    first_json, first_markdown = write_reports(report, tmp_path)
    first_json_content = first_json.read_text(encoding="utf-8")
    first_markdown_content = first_markdown.read_text(encoding="utf-8")

    second_json, second_markdown = write_reports(report, tmp_path)

    assert second_json.read_text(encoding="utf-8") == first_json_content
    assert second_markdown.read_text(encoding="utf-8") == first_markdown_content
    assert json.loads(first_json_content) == report
    assert render_markdown(report) == first_markdown_content

